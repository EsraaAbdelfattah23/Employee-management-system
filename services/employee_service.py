import csv
import os
from typing import List, Tuple, Dict, Any, Optional
from models.employee import Employee
from db.database import Database
from utils.validators import validate_employee_data

# Import libraries for PDF and Excel export
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl

class EmployeeService:

    def __init__(self, database: Database):
    
        self.db = database

    def add_employee(self, employee_data: Dict[str, Any]) -> Tuple[bool, str, Optional[int]]:
        """
        Add a new employee to the database after validation.

        Args:
            employee_data: Dictionary containing employee data

        Returns:
            Tuple of (success, message, employee_id)
        """
        # Validate employee data
        is_valid, error_message = validate_employee_data(employee_data)
        if not is_valid:
            return False, error_message, None

        # Create employee object
        employee = Employee(
            name=employee_data['name'],
            age=employee_data['age'],
            job=employee_data['job'],
            email=employee_data['email'],
            gender=employee_data['gender'],
            phone=employee_data['phone'],
            address=employee_data['address']
        )

        # Insert into database
        employee_id = self.db.insert_employee(employee)
        return True, "Employee added successfully", employee_id

    def update_employee(self, employee_id: int, employee_data: Dict[str, Any]) -> Tuple[bool, str, Optional[int]]:
        """
        Update an existing employee after validation.

        Args:
            employee_id: ID of the employee to update
            employee_data: Dictionary containing updated employee data

        Returns:
            Tuple of (success, message, employee_id)
        """
        # Validate employee data
        is_valid, error_message = validate_employee_data(employee_data)
        if not is_valid:
            return False, error_message, None

        # Create employee object with ID
        employee = Employee(
            id=employee_id,
            name=employee_data['name'],
            age=employee_data['age'],
            job=employee_data['job'],
            email=employee_data['email'],
            gender=employee_data['gender'],
            phone=employee_data['phone'],
            address=employee_data['address']
        )

        try:
            # Update in database
            success = self.db.update_employee(employee)
            if success:
                return True, "Employee updated successfully", employee_id
            else:
                return False, f"Failed to update employee with ID {employee_id}", None
        except Exception as e:
            return False, f"Error updating employee: {str(e)}", None

    def delete_employee(self, employee_id: int) -> Tuple[bool, str, Optional[int]]:
        """
        Delete an employee from the database.

        Args:
            employee_id: ID of the employee to delete

        Returns:
            Tuple of (success, message, employee_id)
        """
        try:
            success = self.db.remove_employee(employee_id)
            if success:
                return True, "Employee deleted successfully", employee_id
            else:
                return False, f"Failed to delete employee with ID {employee_id}", None
        except Exception as e:
            return False, f"Error deleting employee: {str(e)}", None

    def get_all_employees(self) -> Tuple[bool, str, List[Tuple]]:
        """
        Get all employees from the database.

        Returns:
            Tuple of (success, message, employee_data)
        """
        try:
            employees = self.db.fetch_all_employees()
            return True, "Employees retrieved successfully", employees
        except Exception as e:
            return False, f"Error retrieving employees: {str(e)}", []

    def get_employee_by_id(self, employee_id: int) -> Tuple[bool, str, Optional[Employee]]:
        """
        Get an employee by ID.

        Args:
            employee_id: ID of the employee to retrieve
        """
        try:
            employees = self.db.fetch_all_employees()
            for emp in employees:
                if emp[0] == employee_id:
                    return True, "Employee found", Employee.from_tuple(emp)
            return False, f"Employee with ID {employee_id} not found", None
        except Exception as e:
            return False, f"Error retrieving employee: {str(e)}", None

    def export_to_csv(self, filename: str) -> Tuple[bool, str, Optional[str]]:
        """
        Export all employee data to a CSV file.

        Args:
            filename: Path to save the CSV file

        """
        try:
            # Get all employees
            success, _, employees = self.get_all_employees()
            if not success:
                return False, "Failed to retrieve employee data", None

            # Write to CSV file
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # Write header
                writer.writerow(['ID', 'Name', 'Age', 'Job', 'Email', 'Gender', 'Phone', 'Address'])

                # Write data
                writer.writerows(employees)

            return True, "Data exported to CSV successfully", filename
        except Exception as e:
            error_msg = f"Error exporting to CSV: {e}"
            print(error_msg)
            return False, error_msg, None

    def export_to_excel(self, filename: str) -> Tuple[bool, str, Optional[str]]:
        """
        Export all employee data to an Excel file.

        Args:
            filename: Path to save the Excel file
        """
        try:
            # Get all employees
            success, _, employees = self.get_all_employees()
            if not success:
                return False, "Failed to retrieve employee data", None

            # Create a new workbook and select the active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Employees"

            # Write header
            headers = ['ID', 'Name', 'Age', 'Job', 'Email', 'Gender', 'Phone', 'Address']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num).value = header

            # Write data
            for row_num, employee in enumerate(employees, 2):
                for col_num, value in enumerate(employee, 1):
                    sheet.cell(row=row_num, column=col_num).value = value

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            workbook.save(filename)
            return True, "Data exported to Excel successfully", filename
        except Exception as e:
            error_msg = f"Error exporting to Excel: {e}"
            print(error_msg)
            return False, error_msg, None

    def export_to_pdf(self, filename: str) -> Tuple[bool, str, Optional[str]]:
        """
        Export all employee data to a PDF file.

        Args:
            filename: Path to save the PDF file
        """
        try:
            # Get all employees
            success, _, employees = self.get_all_employees()
            if not success:
                return False, "Failed to retrieve employee data", None

            # Create PDF document
            doc = SimpleDocTemplate(filename, pagesize=landscape(letter))
            elements = []

            # Add title
            styles = getSampleStyleSheet()
            title = Paragraph("Employee Management System - Employee Report", styles['Title'])
            elements.append(title)
            elements.append(Paragraph("<br/><br/>", styles['Normal']))

            # Prepare data for table
            headers = ['ID', 'Name', 'Age', 'Job', 'Email', 'Gender', 'Phone', 'Address']
            data = [headers]  # First row is the header
            for employee in employees:
                data.append(employee)

            # Create table
            table = Table(data)

            # Style the table
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ])

            # Add alternating row colors
            for i in range(1, len(data)):
                if i % 2 == 0:
                    style.add('BACKGROUND', (0, i), (-1, i), colors.lightgrey)

            table.setStyle(style)
            elements.append(table)

            # Build PDF
            doc.build(elements)
            return True, "Data exported to PDF successfully", filename
        except Exception as e:
            error_msg = f"Error exporting to PDF: {e}"
            print(error_msg)
            return False, error_msg, None


